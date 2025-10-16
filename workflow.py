import os
import sys
import shutil
import numpy as np
import pandas as pd
from datetime import datetime
import tarfile
sys.path.insert(0,os.path.abspath(os.path.join('dependencies')))
import pyemu
import model_functions
import ies_functions
import pastas_workflow


def _prep(site_name,use_delay,org_dir,num_reals,obs_file,tpl_dir,
          use_focus_weights,use_obs_diff,prepfunc_kwargs,freq,
          start_datetime,verification_window,include_ghb_pars,
          wl_sample,estimate_clay_thickness,
          specified_initial_interbed_state,
          assimilate_all,prefer_less_rebound,
          prefer_less_delayedfuture,include_pastas,
          head_based,pcs0_range,tie_ib_by_layer,run_mod_fxn):


    model_functions.build_model(site_name,prerun=True,use_delay=use_delay,
                                prepfunc_kwargs=prepfunc_kwargs,freq=freq,
                                start_datetime=start_datetime,wl_sample=wl_sample,
                                specified_initial_interbed_state=specified_initial_interbed_state,
                                head_based=head_based)
    # if include_pastas:
    #     model_functions.pastas_get_model(site_name, buffer_miles=2, rfunc="hantush")


    # BUILD PEST INTERFACE
    ies_functions.setup_pst(org_d=org_dir, site_name=site_name, template_ws=tpl_dir,num_reals=num_reals,
        include_ghb_pars=include_ghb_pars,estimate_clay_thickness=estimate_clay_thickness,
                            include_pastas=include_pastas,pcs0_range=pcs0_range,
                            tie_ib_by_layer=tie_ib_by_layer) # todo
    
    # ASSIGN STANDARD DEVIATION TO OBS
    ies_functions.set_obsvals_and_weights(obs_file,t_d=tpl_dir,num_reals=num_reals,
        use_focus_weights=use_focus_weights,use_obs_diff=use_obs_diff,
                                          verification_window=verification_window,
                                          assimilate_all=assimilate_all,
                                          prefer_less_rebound=prefer_less_rebound,
                                          prefer_less_delayedfuture=prefer_less_delayedfuture)

    if run_mod_fxn:
        ies_functions.try_run_modify_pst_fxn(site_name,tpl_dir)
    

def _run(master_dir,tpl_dir,noptmax,num_reals,num_workers,port,usecondor=False):

    ies_functions.prep_for_parallel(b_d=tpl_dir,ct_d=tpl_dir+"_clean",
                          noptmax=noptmax,num_reals=num_reals)

    if usecondor:
        import write_condor
        write_condor.run_condor(template_ws=tpl_dir+"_clean",master_dir=master_dir,num_workers=num_workers,pestpp="ies")
    else:
        ies_functions.run_local(worker_dir=tpl_dir+"_clean",master_dir=master_dir,num_workers=num_workers,pestpp="ies",port=port)


def _plot(master_dir,plot_all_iters=True):
    ies_functions.plot_all(master_dir,plot_all_iters=plot_all_iters)

def analyze_site(site_name,org_dir=None,tpl_dir=None,
                 master_dir=None,obs_file=None,noptmax=10,use_delay=True,
                 num_reals=100,num_workers=20,prep=True,run=True,plot=True,
                 use_focus_weights=True, ht_directions=None,
                 experiment_tag="",use_obs_diff=False, prepfunc_kwargs={},
                 port=4004,freq="Y",start_datetime=None,verification_window=None,
                 include_ghb_pars=True,wl_sample="mean",run_scenarios=True,plot_scens=False,
                 scenario_subset=None,estimate_clay_thickness=False,usecondor=False,
                 scenario_tag="",scenariofiles=[],run_prior_scenarios=False,
                 export_scenario_basereals=True,morris_scenario_name=None,
                 specified_initial_interbed_state=True,xfer_from_m_d=None,
                 assimilate_all=False,prefer_less_rebound=1e-20,
                 prefer_less_delayedfuture=False,
                 include_pastas=False,head_based=False,pcs0_range=200,
                 tie_ib_by_layer=True,run_mod_fxn=True):

    if verification_window is not None:
        assert len(verification_window) == 2

    if org_dir is None:
        org_dir = os.path.join(site_name, 'model_ws', "historical")

    if obs_file is None:
        
        obs_file = os.path.join(site_name,"source_data","{0}_sub_data.csv".format(site_name))
    
    if master_dir is None:
        master_dir = os.path.join(site_name,"master_ies")

        #delay/nodelay tag
        if use_delay:
            master_dir += "_delay"
        else:
            master_dir += "_nodelay"

        # frequency tag
        master_dir += "_{0}".format(freq.lower())

        # focus/nofocus tage
        if use_focus_weights:
            master_dir += "_focus"
        else:
            master_dir += "_nofocus"

        # specified_initial_interbed_state tag
        if specified_initial_interbed_state:
            master_dir += "_specstate"
        else:
            master_dir += "_nospecstate"

        if prefer_less_rebound > 1e-10:
            master_dir += "_lessrbnd"

        if prefer_less_delayedfuture:
            master_dir += "_lessfuture"

        if head_based:
            master_dir += "_headbased"
        else:
            master_dir += "_stressbased"
        if tie_ib_by_layer == "all":
            master_dir += "_ibtieall"
        elif tie_ib_by_layer is True:
            master_dir += "_ibbylayer"
        else:
            master_dir += "_noibbylayer"

        if run_mod_fxn:
            master_dir += "_modfxn"
        else:
            master_dir += "_nomodfxn"


        # exp tag
        assert len(experiment_tag) > 0, "set an experiment tag!"
        # if len(experiment_tag) == 0:
        #     experiment_tag = datetime.now().strftime("%Y%m%d_%H%M%S")
        master_dir += experiment_tag

    else:
        master_dir = os.path.join(site_name, master_dir)
    print(f"master dir: {master_dir}")

    if tpl_dir is None:
        tpl_dir = master_dir.replace("master_","template_")
    print(f"template dir: {tpl_dir}")

    if xfer_from_m_d is not None:
        assert os.path.exists(xfer_from_m_d)
        print(f"transfer master dir from {xfer_from_m_d}")

    if prep:
        _prep(site_name,use_delay,org_dir,num_reals,obs_file,tpl_dir,
          use_focus_weights,use_obs_diff,prepfunc_kwargs,freq,
          start_datetime,verification_window,include_ghb_pars,
          wl_sample,estimate_clay_thickness,
          specified_initial_interbed_state,
          assimilate_all,prefer_less_rebound,
          prefer_less_delayedfuture,include_pastas,
          head_based,pcs0_range,tie_ib_by_layer,run_mod_fxn)

        if xfer_from_m_d is not None:
            ies_functions.xfer_from_m_d(tpl_dir,xfer_from_m_d)

    if run:
       ies_m_d = _run(master_dir, tpl_dir, noptmax, num_reals, num_workers, port, usecondor=usecondor)

       #ies_functions.export_realizations_to_dirs(tpl_dir,master_dir,
       # real_name_tag="",noptmax=None,just_mf6=True)

    if plot:
        _plot(master_dir,plot_all_iters=True)


    if morris_scenario_name == "baseline":
        morris_m_d = ies_functions.run_morris(org_t_d=tpl_dir,
                                              num_workers=num_workers,
                                              scenario_name=None,plusplus_kwargs={"gsa_morris_p":10,"gsa_morris_r":10})
        ies_functions.plot_morris_delaydif_summary(morris_m_d)
        if os.path.exists(master_dir):
            ies_functions.plot_morris_delaydif_summary2(master_dir,morris_m_d)

    if run_scenarios:

        assert os.path.exists(master_dir), f"{master_dir} not found"
        prep_CH_scen_Input(site_name)
        
        # if len(scenario_tag) == 0:
        #     scenario_tag = "_scenarios"

        # same scenario tags for everyone - don't change
        scenario_dict = {
            "DWR_scenarios": "scenario_data",
            "CH_forecast": "CH_forecast",
            "2015_scenario": "scenario_data_2015",
            "MTNoMO_scenario": "MTNoMO_scenario_data",
            "HCMT_scenario": "HCMTscenario",
            "TH_scenarios":"TH_scenario_data",
            "TH_reduction_scenario": "TH_reduction_scenario_data"
        }

        # testing_scanlayers(scenario_dict,kdict)
        # testing_prep_scen_csv(scenario_dict)

        # filter scenario_dict based on list of scenariofiles
        print(list(scenario_dict.keys()))
        print(scenariofiles)
        assert [s in scenario_dict.keys() for s in scenariofiles]
        scenario_dict_temp = scenario_dict.copy()
        for k in scenario_dict.keys():
            if k not in scenariofiles:
                scenario_dict_temp.pop(k)
        assert len(scenario_dict_temp) > 0

        # add additional string to scenario_tag
        if len(scenario_tag) > 0:
            scenario_dict_temp = {f"{k}_{scenario_tag}": v for k, v in scenario_dict_temp.items()}
        scenario_dict_temp = {f"_{k}": v for k, v in scenario_dict_temp.items()}

        # loop through scenarios
        for scenario_tag_final, scenariofile in scenario_dict_temp.items():

            # prep the scenario csv
            prep_scenario_csv_general(site_name, scenariofile)

            # run the scenarios
            tpl_dir_new = tpl_dir + "_scenarios"
            master_dir_new = master_dir + scenario_tag_final
            ies_functions.run_scenarios(site_name, t_d=os.path.join(tpl_dir), new_t_d=tpl_dir_new,
                                        m_d=os.path.join(master_dir), new_m_d=master_dir_new,
                                        subset_size=scenario_subset,
                                        usecondor=usecondor,num_workers=num_workers,scenario_tag=scenario_tag_final,
                                        run_prior_scenarios=run_prior_scenarios,
                                        scenariofile=scenariofile, port=port)
            ies_functions.plot_en_compaction_scenarios(m_d=os.path.join(master_dir+scenario_tag_final))
            # ies_functions.plot_en_subsidence_scenarios(site_name, m_d=os.path.join(master_dir), new_m_d=master_dir_new)
            if run_prior_scenarios:
                ies_functions.plot_en_compaction_scenarios(m_d=os.path.join(master_dir+scenario_tag_final+"_prior"),
                                                           bayes_stance="pr")
            if export_scenario_basereals:
                ies_functions.export_realizations_to_dirs(t_d=tpl_dir, m_d=master_dir+scenario_tag_final,
                             real_name_tag="real:base",noptmax=None,just_mf6=True,scenario_tag=scenario_tag_final)
            # if morris_scenario_name is not None:
            #     #try:
            #         morris_m_d = ies_functions.run_morris(org_t_d=tpl_dir_new,
            #                                               num_workers=num_workers,scenario_name=morris_scenario_name)
            #         ies_functions.plot_morris_summuary(morris_m_d)
            #     #except Exception as e:
            #     #    print("error running scenario analysis:{0}".format(str(e))

    if plot_scens:
        try:
            ies_functions.plot_scenarios(site_name=site_name,
                                         m_d=master_dir)
        except:
            pass

    if ht_directions is not None:
        if not isinstance(ht_directions,list):
            ht_directions = [ht_directions]
        for ht_direction in ht_directions:
            tpl_dir_ht = tpl_dir + "_ht_"+ht_direction
            #prep_for_hypoth_test(ht_direction,org_t_d,org_m_d,new_t_d=None,post_noptmax=None)
            ies_functions.prep_for_hypoth_test(ht_direction,tpl_dir,master_dir,new_t_d=tpl_dir_ht,post_noptmax=None)
            master_dir_ht = master_dir + "_ht_" + ht_direction
            _run(master_dir, tpl_dir, noptmax, num_reals, num_workers, port, usecondor=usecondor)
            _plot(master_dir_ht,plot_all_iters=True)
            ies_functions.plot_compare_ht(master_dir,master_dir_ht,ht_oname=None)

    return master_dir

def try_analysis_for_sites(sites,num_workers,num_reals,noptmax,use_focus_weights):
    errors = []
    for site_name in sites:
        try:
            analyze_site(site_name,num_reals=num_reals,noptmax=noptmax,num_workers=num_workers,run=True,prep=True,plot=True,
                use_focus_weights=use_focus_weights)
        except Exception as e:
            errors.append("error analyzing delay form for site '{0}': {1}".format(site_name,str(e)))
        try:
            analyze_site(site_name,num_reals=num_reals,noptmax=noptmax,num_workers=num_workers,run=True,prep=True,plot=True,
                use_delay=False,use_focus_weights=use_focus_weights)
        except Exception as e:
            errors.append("error analyzing no-delay form for site '{0}': {1}".format(site_name,str(e)))
    if len(errors) > 0:
        raise Exception("errors during analyses: {0}".format('\n\n'.join(errors)))

def gather_pdfs(sites, dest_d = "pdfs",combine_to_one=False):
    os.makedirs(dest_d, exist_ok=True)
    for site in sites:
        m_ds = []
        dd = [os.path.join(site, dd) for dd in os.listdir(os.path.join(site)) if
              os.path.isdir(os.path.join(site, dd)) and dd.startswith("master_ies")]
        m_ds.extend(dd)

        for d in m_ds:
            csvs = [f for f in os.listdir(d) if f.endswith(".pdf")]
            dst_dcombined = os.path.join(dest_d, d)
            dst_d = os.path.join(d, 'pdfs')
            os.makedirs(dst_d, exist_ok=True)
            os.makedirs(dst_dcombined, exist_ok=True)
            os.makedirs(dst_d, exist_ok=True)
            for csv in csvs:
                src_f = os.path.join(d, csv)
                dest_f = os.path.join(dst_d, csv)
                dest_fcombined = os.path.join(dst_dcombined, csv)
                print(src_f, dest_f)
                shutil.copy2(src_f, dest_f)
                shutil.copy2(src_f, dest_fcombined)

def gather_csvs(sites, dest_d = "csvs",combine_to_one=False):
    # if os.path.exists(dest_d):
    #     shutil.rmtree(dest_d)
    os.makedirs(dest_d, exist_ok=True)
    for site in sites:
        m_ds = []
        dd = [os.path.join(site,dd) for dd in os.listdir(os.path.join(site)) if os.path.isdir(os.path.join(site,dd)) and dd.startswith("master_ies")]
        m_ds.extend(dd)

        for d in m_ds:
            csvs = [f for f in os.listdir(d) if f.endswith(".csv")]
            dst_dcombined = os.path.join(dest_d, d)
            dst_d = os.path.join(d, 'csvs')
            os.makedirs(dst_d, exist_ok=True)
            os.makedirs(dst_dcombined, exist_ok=True)
            for csv in csvs:
                src_f = os.path.join(d,csv)
                dest_f = os.path.join(dst_d, csv)
                dest_fcombined = os.path.join(dst_dcombined, csv)
                print(src_f,dest_f)
                shutil.copy2(src_f,dest_f)
                shutil.copy2(src_f, dest_fcombined)

def gather_pngs(sites, dest_d = "figures",combine_to_one=False):
    os.makedirs(dest_d, exist_ok=True)
    for site in sites:
        m_ds = []
        dd = [os.path.join(site,dd) for dd in os.listdir(os.path.join(site)) if os.path.isdir(os.path.join(site,dd)) and dd.startswith("master_ies")]
        m_ds.extend(dd)

        for d in m_ds:
            pngs = [f for f in os.listdir(d) if f.endswith(".png")]
            sd = os.path.join(d,"figures")
            if os.path.exists(sd):
                ppngs = [os.path.join("figures",f) for f in os.listdir(sd) if f.endswith(".png")]
                pngs.extend(ppngs)
            dst_d = os.path.join(d, 'figures')
            dst_dcombined = os.path.join(dest_d, d)
            os.makedirs(dst_d, exist_ok=True)
            os.makedirs(dst_dcombined, exist_ok=True)
            for png in pngs:
                if 'figures' not in png:
                    src_f = os.path.join(d, png)
                    dest_f = os.path.join(dst_d, os.path.split(png)[1])
                    print(png,src_f, dest_f)
                    shutil.copy2(src_f, dest_f)
                    shutil.copy2(src_f, dest_fcombined)

def get_water_levels_minrolled(tsdf):
    # .rolling(): apply rolling window on dataframe with:
    #     window size = 90 (days)
    #     center=True: centered on the current value
    #     min_periods=1: at least 1 non-NaN value required within the window for computation
    # .min(): then get the minimum value within each rolling window
    # .dropna(): then drop nans
    tsdf = tsdf.rolling(90,center=True,min_periods=1).min().dropna()
    return tsdf


def gather_results(site_names,tag=""):

    r_d = "results" + tag
    if not os.path.exists(r_d):
        #shutil.rmtree(r_d)
        os.makedirs(r_d)

    for site_name in site_names:
        m_ds = [os.path.join(site_name,m_d) for m_d in os.listdir(site_name) if os.path.isdir(os.path.join(site_name,m_d)) and "master" in m_d.lower()]
        for m_d in m_ds:
            phi_file = os.path.join(m_d,"pest.phi.actual.csv")
            mdate = os.path.getmtime(phi_file)
            mdate = datetime.fromtimestamp(mdate)
            mdate = mdate.strftime("%Y%m%d_%H%M%S")
            dest_d = os.path.join(r_d,site_name,mdate+"_"+site_name+"_"+os.path.split(m_d)[1])
            if os.path.exists(dest_d):
                shutil.rmtree(dest_d)
            os.makedirs(dest_d)
            pst_file = os.path.join(m_d,"pest.pst")
            if not os.path.exists(pst_file):
                continue
            pst = pyemu.Pst(pst_file)
            pst.write(os.path.join(dest_d,"pest.pst"),version=2)
            for f in ["pest.rec","pest.rmr","pest.log"]:
                if os.path.exists(os.path.join(m_d,f)):
                    shutil.copy2(os.path.join(m_d,f),os.path.join(dest_d,f))
            for d in ["source_data","processed_data"]:
                src_d = os.path.join(m_d,d)
                if os.path.exists(src_d):
                    shutil.copytree(src_d,os.path.join(dest_d,d))

            phidf = pd.read_csv(phi_file)
            files = [f for f in os.listdir(m_d) if f.endswith("base.par")]
            for f in files:
                shutil.copy2(os.path.join(m_d,f),os.path.join(dest_d,f))
            files = [f for f in os.listdir(m_d) if f.endswith(".py.txt")]
            for f in files:
                shutil.copy2(os.path.join(m_d,f),os.path.join(dest_d,f))
            files = [f for f in os.listdir(m_d) if f.startswith("meandata") or f.startswith("basedata")]
            for f in files:
                shutil.copy2(os.path.join(m_d,f),os.path.join(dest_d,f))
            files = [f for f in os.listdir(m_d) if f.startswith("pest.phi") and f.endswith(".csv")]
            for f in files:
                shutil.copy2(os.path.join(m_d,f),os.path.join(dest_d,f))
            files = [f for f in os.listdir(m_d) if f.startswith("critical_results") and f.endswith(".csv")]
            for f in files:
                shutil.copy2(os.path.join(m_d,f),os.path.join(dest_d,f))
            

            itrs = [0,phidf.iteration.max()]
            for i in itrs:
                oe_file = "pest.{0}.obs.jcb".format(i)
                if os.path.exists(os.path.join(m_d,oe_file)):
                    shutil.copy2(os.path.join(m_d,oe_file),
                        os.path.join(dest_d,oe_file))
                pe_file = "pest.{0}.par.jcb".format(i)
                if os.path.exists(os.path.join(m_d,pe_file)):
                    shutil.copy2(os.path.join(m_d,pe_file),
                        os.path.join(dest_d,pe_file))

                
                files = [f for f in os.listdir(m_d) if str(i) in f and f.endswith(".pdf")]
                for f in files:
                    shutil.copy2(os.path.join(m_d,f),os.path.join(dest_d,f))

def prep_scenario_csv_general(location, scenariofile):

    print(f"prep csv for {scenariofile}")

    lith = pd.read_csv(os.path.join(location, "source_data", f"{location}_lithology.csv")).dropna(axis=1, how="all").dropna(axis=0)
    lith.Aquifer = lith.Aquifer.str.lower()
    lith_lays = [s.replace(" ","").replace("aquifer","").replace("plioscene","pliocene").replace("composite","corcoran") for s in lith.Aquifer.unique()]
    lith_dict = dict(zip(lith_lays, range(len(lith_lays))))
    lith_nlay = len(lith.Aquifer.unique())
    print(f"{lith_nlay} layers in lithology file: {lith_dict}")

    # load scenario data and get layer names
    scen_df = pd.read_excel(os.path.join(location, "source_data", f"{location}_{scenariofile}.xlsx"), header=None)
    scen_lays = [lay for lay in scen_df.iloc[0, 1:].str.lower().unique().tolist() if pd.notna(lay)]
    scen_lays = [s.replace("plioscene","pliocene").replace("composite","corcoran") for s in scen_lays] # composite=corcoran. todo cehck with John Ellis

    # get kvals: the zero-based cols for scenario data
    if scenariofile == "CH_forecast": #  {"layer_1": 0, "layer_2": 1, "layer_3": 2, "layer_4": 3, "layer_5": 4}
        kvals = [int(s.replace("layer_",""))-1 for s in scen_lays]
    else:
        kvals = [lith_dict[s] for s in scen_lays]
    print(f"{len(kvals)} layers in scenario file: {dict(zip(scen_lays,kvals))}")

    # get scenario_js: zero-based model layers that the scenario data columns map to
    scen_nlay = len(kvals)
    no_scen = int((scen_df.shape[1] - 1) / len(kvals))  # number of scenarios = number of sublists
    cols = list(range(1, scen_df.shape[1]))  # list of column numbers
    scenario_js = [cols[i:i + scen_nlay] for i in range(0, no_scen * scen_nlay, scen_nlay)]  # split into sublists

    # try to make some better scenario names
    scen_names = [scen_df.iloc[1, sjs[0]] for sjs in scenario_js]
    print(f"scen names are {scen_names}")
    scen_names = [sn.lower().replace(" ", "-").replace("_", "-").replace("(", "").replace(")", "") for sn in scen_names]

    # the zero based col for scenario dates
    date_j = 0
    # zero-based row where dates/data start
    start_i = 3
    dates = pd.to_datetime(scen_df.iloc[start_i:, date_j])

    scen_dict = {}
    for name, js in zip(scen_names, scenario_js):
        vals = scen_df.iloc[start_i:, js].values
        assert vals.shape[1] == len(kvals)
        # skip scenarios where rows are filled with nans
        if all(x is np.nan for sublist in vals for x in sublist):
            continue
        else:
            for k, v in zip(kvals, vals.transpose()):
                sname = name + "_k:{0}".format(k)
                assert sname not in scen_dict
                scen_dict[sname] = v

    scen_df = pd.DataFrame(scen_dict, index=dates)
    scen_df.index.name = "datetime"
    scen_df = scen_df.ffill() # fill nan values by propagating the last valid observation to next valid
    scen_df.dropna(axis=0, how="all", inplace=True)

    if scenariofile == "CH_forecast":

        # load obs data to get layers in which there are observations
        obsdata_df = pd.read_csv(os.path.join(location, "source_data", f"{location}_obs_data.csv"), index_col=0)
        lays_obs = obsdata_df["Aquifer"].str.lower().dropna().unique().tolist()
        lays_obs = [s.replace("plioscene","pliocene").replace("composite","corcoran").replace(" ", "") for s in lays_obs]
        kvals_obs = [lith_dict[s] for s in lays_obs]
        print(f"observations in: {dict(zip(lays_obs, kvals_obs))}")

        # only keep layers in which there are observations
        kvals_ch = [k for k in kvals if k in kvals_obs]
        print(f"{len(kvals_ch)} layers in scenario file with observations: {kvals_ch}")

        # only keep scenario names that contain 'k:{i}' for i in kvals_ch
        scen_names_ch = [s for s in scen_df.columns if any(f'k:{i}' in s for i in kvals_ch)]
        scen_df = scen_df[scen_names_ch]

        # load scenario data and get layer names
        # scen_dfcheck = pd.read_excel(os.path.join(location, "source_data", f"{location}_scenario_data.xlsx"), header=None)
        # layscheck = scen_dfcheck.iloc[0, 1:].str.lower().replace("plioscene","pliocene").unique().tolist()
        # assert all([s in lith_dict.keys() for s in layscheck]), f"an aquifer name in scenario_data is not recorded in kdict"

    # if ("composite" in scen_lays) or (scenariofile == "CH_forecast" and "composite" in layscheck):
    #
    #     # don't track composite layer in scenarios
    #     if ("composite" in scen_lays):
    #         lays_nocomp = [s for s in scen_lays if s != "composite"]
    #     else:
    #         lays_nocomp = [s for s in scen_lays if s != "layer_2"] # "composite" layer = layer 2 (see kdict)
    #     assert len(lays_nocomp) < len(scen_lays)
    #     kvals_nocomp = [kdict[s] for s in lays_nocomp]
    #     # only keep scenario names that contain 'k:{i}' for i in kvals_new
    #     scen_names_ch = [s for s in scen_df.columns if any(f'k:{i}' in s for i in kvals_nocomp)]
    #     scen_df = scen_df[scen_names_ch]

    print(f"columns names: {scen_df.columns.tolist()}")
    if not os.path.exists(os.path.join(location, "processed_data")):
        os.mkdir(os.path.join(location, "processed_data"))
    scen_df.to_csv(os.path.join(location, "processed_data", f"{location}.scenarios.csv"))

def testing_scanlayers(scenario_dict):
    sites = [subdir for subdir in next(os.walk('.'))[1] if subdir not in
             [".git", ".idea", "__pycache__", "bin", "CSUB_example", "CVHM2", "dependencies", "etc", "gis", "log",
              "temp", "THmodel", "Tule"]]

    testdic = scenario_dict.copy()
    testdic["obs_data"] = "obs_data"
    testdic["par_data"] = "par_data"
    testdic["lithology"] = "lithology"

    lay_dic = {}
    lay_dic_u = {}
    for tag, file in testdic.items():
        print(tag)
        lays_u = []
        lay_dic_temp = {}
        for location in sites:
            # check for file
            if (file == "obs_data") or (file == "lithology"):
                fpath = os.path.join(location, "source_data", f"{location}_{file}.csv")
            else:
                fpath = os.path.join(location, "source_data", f"{location}_{file}.xlsx")

            if not os.path.exists(fpath):
                print(f"no file for {location}")

            else:
                # if file exists, read lay names
                if file == "obs_data":
                    lays = pd.read_csv(fpath, index_col=0)["Aquifer"].str.lower().dropna().unique().tolist()
                elif file == "lithology":
                    lays = pd.read_csv(fpath).dropna(axis=1,how="all").dropna(axis=0).Aquifer.str.lower().unique()
                    lays = [s.replace(" ", "").replace("aquifer", "") for s in lays]
                elif file == "par_data":
                    lays = pd.read_excel(fpath).columns.str.lower().tolist()
                    # truncate to exclude the first occurrence of 'parameter' and keep values after it
                    lays = lays[lays.index('parameter') + 1:]
                    # remove all values after the first occurrence of a value containing 'unnamed'
                    if any("unnamed" in s for s in lays):
                        lays = [x for x in lays if not ('unnamed' in x)][0:lays.index(next(x for x in lays if 'unnamed' in x))]
                else:
                    lays = pd.read_excel(fpath, header=None).iloc[0, 1:].str.lower().unique().tolist()
                lays = [lay for lay in lays if pd.notna(lay)]
                lays = [s.strip() for s in lays]
                lay_dic_temp[location] = lays
                lays_u.extend([s for s in lays if s not in lays_u])
                # assert all(s in kdict for s in lays_u), f"an aquifer name in {file} is not recorded in kdict"

        lay_dic[tag] = lay_dic_temp
        lay_dic_u[tag] = lays_u

    with open(os.path.join(".", "testing_scanlayers.txt"), 'w') as f:
        for key, value in lay_dic.items():
            f.write(f"-------------------------------- {key}\n")
            for k, v in value.items():
                f.write(f"{k}: " + ", ".join(v) + "\n")
            f.write(f"UNIQUE: " + ", ".join(lay_dic_u[key]) + "\n")

def testing_prep_scen_csv(scenario_dict):

    sites = [subdir for subdir in next(os.walk('.'))[1] if subdir not in
             [".git", ".idea", "__pycache__", "bin", "CSUB_example", "CVHM2", "dependencies", "etc", "gis", "log",
              "temp", "THmodel", "Tule"]]

    with open(os.path.join(".", "testing_prep_scen_csv.txt"), 'w') as f:
        for site in sites:
            f.write(f"-------------------------------- {site}\n")
            for scenario_tag, scenariofile in scenario_dict.items():
                    if os.path.exists(os.path.join(site, "source_data", f"{site}_{scenariofile}.xlsx")):
                        # prep the scenario csv
                        try:
                            # Redirect print statements to the file
                            original_stdout = sys.stdout  # Save the original stdout
                            sys.stdout = f  # Redirect to file
                            prep_scenario_csv_general(site, scenariofile)
                            # Restore the original stdout
                            sys.stdout = original_stdout
                            f.write(f"SUCCESS for {scenariofile}\n\n")
                        except Exception as e:
                            # Restore stdout and log failure
                            sys.stdout = original_stdout
                            f.write(f"FAILURE for {scenariofile}: {e}\n\n")
    # Reset sys.stdout to default
    sys.stdout = sys.__stdout__

def prep_CH_scen_Input(site):


    m_ds = []
    dd = [os.path.join(site, dd) for dd in os.listdir(os.path.join(site)) if
          os.path.isdir(os.path.join(site, dd)) and dd.startswith("master_ies")]
    m_ds.extend(dd)

    for d in m_ds:
        if 'scenarios' not in d:

            lith = pd.read_csv(os.path.join(site, "source_data", f"{site}_lithology.csv")).dropna(axis=1,
                                                                                                          how="all").dropna(
                axis=0)
            lith.Aquifer = lith.Aquifer.str.lower()
            lith_lays = [
                s.replace(" ", "").replace("aquifer", "").replace("plioscene", "pliocene").replace("composite", "corcoran")
                for s in lith.Aquifer.unique()]
            lith_dict = dict(zip(lith_lays, range(1, len(lith_lays)+1)))
            lith_nlay = len(lith.Aquifer.unique())
            print(f"{lith_nlay} layers in lithology file: {lith_dict}")



            from pathlib import Path
            folders = os.path.normpath(d).split(os.path.sep)
            noptmax = len([s for s in os.listdir(os.path.join(site, f'{folders[1]}')) if s.startswith('mean_delayib_LowestGWL')])
            CHInput = f'mean_delayib_LowestGWL_{noptmax-1}.csv'

            if CHInput in os.listdir(d):

                df = pd.read_csv(os.path.join(site, f'{folders[1]}',f'{CHInput}'), header=None, delimiter=',')
                df.columns = df.iloc[0]
                df = df[1:]
                # Convert to datetime format
                df['datetime'] = pd.to_datetime(df['datetime'])

                df_measured = pd.read_csv(os.path.join(site, "processed_data", f"{site}.ts_data.csv"), header=None, delimiter=',')
                df_measured = df_measured[1:]
                df_measured.columns = ['Date', 'Alt', 'interpolated', 'Aquifer', 'klayer']
                for Aq in df_measured['Aquifer'].unique():
                    lay = lith_dict[Aq.replace(" ", "").lower()]
                    df_measured.loc[df_measured['Aquifer'] == Aq,'Aquifer'] = f'layer_{lay}'

                df_measured = df_measured.set_index(pd.to_datetime(df_measured['Date']),
                                                    drop=True)
                # if site in ['438.939', '376.676', '341.804', 'Y88', 'H201']:
                #     df_measured = df_measured.loc[df_measured['Aquifer'].isin(['Upper', 'Lower']), :]

                # Drop CH data if there is no measured equivalent
                for col in df.keys():
                    if 'layer' in col:
                        if col not in df_measured['Aquifer'].unique():
                            df = df.drop(col, axis=1)

                # Also drop the measured data if there is no CH equivalent
                for aquifer in df_measured['Aquifer'].unique():
                    if aquifer not in df.keys():
                        df_measured = df_measured.loc[df_measured['Aquifer'] != aquifer]

                # Grab the layer key for later use
                layer_keys = list(df_measured['Aquifer'].unique())

                # Grab the keys we want to duplicate the scenarios for, don't want to duplicate datetimes
                keys = df.keys().drop('datetime')

                for c in df.keys():
                    if c.startswith('lay'):
                        df[c] = df[c].astype(float)

                # ----------------------------
                # Create +20 and +50 scenarios
                # ----------------------------
                for max_increase in [20, 50]:
                    for layer in keys:
                        # Create a new column for the scenario
                        scenario_col = f'{layer}_{max_increase}'
                        df[scenario_col] = df[layer] + max_increase  # Start with original data

                df_scenario = df.copy()
                df_scenario = df_scenario.set_index('datetime')

                # Merge the scenario with the measured data
                measured_data = pd.DataFrame(
                    index=df_measured.loc[df_measured['Aquifer'] == df_measured['Aquifer'].unique()[0]].index)
                for aquifer in df_measured['Aquifer'].unique():
                    df = pd.DataFrame({f'{aquifer}_measured': df_measured.loc[df_measured['Aquifer'] == aquifer, 'interpolated']})
                    measured_data = pd.merge(measured_data, df,
                                             left_index=True,
                                             right_index=True,
                                             how='outer')

                # Linearly Interpolate
                # measured_data = measured_data.resample('D').interpolate()

                # Start of scenario data
                measured_data = measured_data[measured_data.index < pd.to_datetime('03/01/2024')]

                # Create the final DataFrame that will hold combined data
                final_df = df_scenario.copy()

                # Define the rate of increase or decrease in ft/day (convert from 2.85 ft/month)
                rate_per_day = 2.85 / 30

                # Loop through each layer to create the interpolated values column
                for layer in df_scenario.keys():
                    # Extract the last measured value and date for the layer
                    layer_measured = layer[:7]
                    last_measured = df_measured[df_measured['Aquifer'] == layer_measured].sort_index().iloc[-1]
                    last_measured_date = last_measured.name  # Index holds the datetime
                    last_measured_value = last_measured['interpolated']  # 'Alt' column holds the measured value

                    # Get the scenario value at the final time step
                    final_scenario_value = df_scenario[layer].iloc[-1]

                    # Determine rate direction
                    rate_sign = 1 if float(last_measured_value) < float(final_scenario_value) else -1
                    rate_per_day = rate_sign * (2.85 / 30)

                    flag = False

                    # Interpolate from the last measured date to the end of the scenario data
                    for date in final_df.index:
                        if date <= last_measured_date:
                            # Before the last measured date, keep NaN or use original measured data
                            continue

                        # Calculate the interpolated value
                        days_since_last_measured = (date - last_measured_date).days

                        # Get the scenario value at this date
                        scenario_value = final_df.at[date, layer]

                        interp_value = float(last_measured_value) + (days_since_last_measured * float(rate_per_day))
                        # Set the final value based on the condition
                        if (rate_per_day > 0 and interp_value >= scenario_value) or (
                                rate_per_day < 0 and interp_value <= scenario_value):
                            if not flag:
                                final_df.at[date, f'{layer}_final'] = scenario_value
                                val = scenario_value
                                flag = True
                            else:
                                final_df.at[date, f'{layer}_final'] = val
                        else:
                            final_df.at[date, f'{layer}_final'] = interp_value

                # ---------------------------------------------
                # Create the df for exporting data and plotting
                # ---------------------------------------------
                final_keys = [x for x in final_df.keys() if 'final' in x]
                for i, ch_val in enumerate(['t', '20', '50']):
                    if ch_val == 't':
                        _keys = [f"{x}_final" for x in layer_keys]
                        t = final_df[_keys]
                        col_names = [f"{x}_measured" for x in layer_keys]
                        t.columns = col_names
                        t = t[t.index > pd.to_datetime('03/01/2024')]
                        df1 = pd.concat([measured_data, t], axis=0)
                    else:
                        _keys = [x for x in final_keys if ch_val in x]
                        t = final_df[_keys]
                        col_names = [f"{x}_measured" for x in layer_keys]
                        t.columns = col_names
                        t = t[t.index > pd.to_datetime('03/01/2024')]
                        df2 = pd.concat([measured_data, t], axis=0)
                        if i == 1:
                            results = pd.concat([df1, df2],
                                                axis=1)
                        else:
                            results = pd.concat([results, df2],
                                                axis=1)

                results_columns = [x for x in layer_keys]
                [results_columns.append(f"{x}_20") for x in layer_keys]
                [results_columns.append(f"{x}_50") for x in layer_keys]
                results.columns = results_columns
                results = results[~results.index.duplicated(keep='first')]
                df = results.copy()

                # -------------------------------------------------------
                # Create the multi-index df that matches Excel formatting
                # -------------------------------------------------------
                data = {}
                k = 0
                for scenario in ['Mean CH', 'CH +20 ft', 'CH +50 ft']:
                    for idx, key in enumerate(keys):
                        data[(f'{key}', f'{scenario}', 'WL_alt')] = df[df.columns[idx + k]].values
                    k += len(keys)
                index = pd.MultiIndex.from_tuples(data.keys(), names=['Layer', 'Scenario', 'Date'])
                df_save = pd.DataFrame(data, columns=index)
                dates = df.index
                dates.name = ''
                df_save.index = dates

                # -------------------------------------------------------------------------------
                # Create the new directory to save files to if it does not exist, export to Excel
                # -------------------------------------------------------------------------------
                # Catch a special case for Q288
                if site == 'Q288.mean':
                    site = 'Q288'
                elif site == 'GWM':
                    site = 'GWM-14'

                # Write the data as excel
                df = df.round(1)
                df_save.to_excel(os.path.join(site, 'source_data', f'{site}_CH_forecast.xlsx'),
                                 merge_cells=True,
                                 )

                # ----------------------------------------------------
                # Load and write some formatting to the Excel workbook
                # Merging second row columns, dropping row 4
                # ----------------------------------------------------
                from openpyxl import load_workbook
                wb = load_workbook(os.path.join(site, 'source_data', f'{site}_CH_forecast.xlsx'))
                ws = wb.active
                # Merge header columns
                if len(keys) == 2: #2 layers
                    ws.merge_cells('B2:C2')
                    ws.merge_cells('D2:E2')
                    ws.merge_cells('F2:G2')
                elif len(keys) == 4: #4 layers
                    ws.merge_cells('B2:E2')
                    ws.merge_cells('F2:I2')
                    ws.merge_cells('J2:M2')
                elif len(keys) == 1: #1 layer
                    continue
                elif len(keys) == 3:
                    ws.merge_cells('B2:D2')
                    ws.merge_cells('E2:G2')
                    ws.merge_cells('H2:J2')
                # # Delete row 4
                ws.delete_rows(4)
                wb.save(os.path.join(site, 'source_data', f'{site}_CH_forecast.xlsx'))
                wb.close()
    print(f"Create CH Scenario Input file for site {site}")



def make_tarfile(output_filename, source_dir):
    print("...making archive",output_filename)
    with tarfile.open(output_filename, "w:gz") as tar:
        tar.add(source_dir, arcname=os.path.basename(source_dir))


def make_archive(site,tag=None):
    arc_d = "archive"
    m_ds = [os.path.join(site,m_d) for m_d in os.listdir(site) if os.path.isdir(os.path.join(site,m_d)) 
            and m_d.startswith("master")]
    if tag is not None:
        m_ds = [m_d for m_d in m_ds if tag in m_d]
    print("MASTER DIRS: ",m_ds)
    for m_d in m_ds:
        t_d = m_d.replace("master","template") + "_clean"
        if not os.path.exists(t_d):
            print("WARNING: clean t_d '{0}' not found for m_d '{1}'".format(t_d,m_d))
            continue
        a_d = os.path.join(arc_d,site)
        if not os.path.exists(a_d):
            os.makedirs(a_d)
        a_file = os.path.join(a_d,m_d.split(os.path.sep)[-1])+".{0}.tar.gz".format(datetime.now().strftime("%Y%m%d.%H%M%S"))
        if os.path.exists(a_file):
            print("WARNING: archive file '{0}' already exists, removing...")
            os.remove(a_file)
        tt_d = t_d + "_temp"
        if os.path.exists(tt_d):
            shutil.rmtree(tt_d)
        shutil.copytree(t_d,tt_d)
        pst = pyemu.Pst(os.path.join(m_d,"pest.pst"))
        pst.write(os.path.join(tt_d,"pest.pst"),version=2)

        phidf = pd.read_csv(os.path.join(m_d,"pest.phi.actual.csv"))
        pe_file = "pest.{0}.par.jcb".format(phidf.iteration.max())
        shutil.copy2(os.path.join(m_d,pe_file),os.path.join(tt_d,pe_file))
        shutil.copy2(os.path.join(m_d,pst.pestpp_options["ies_par_en"]),os.path.join(tt_d,pst.pestpp_options["ies_par_en"]))    
        shutil.copy2(os.path.join(m_d,"pest.rec"),os.path.join(tt_d,"pest.rec"))
        make_tarfile(a_file,tt_d)

def make_archives(sites,tag=None):
    for site in sites:
        make_archive(site,tag=tag)


if __name__ == "__main__":
    
    sites = ["J88","Q288","T88","OCTOL"]
    num_workers = 35
    num_reals = 1000
    noptmax = 5
    use_focus_weights = False
    local=True
    #try_analysis_for_sites(site_names,num_workers,num_reals,noptmax,use_focus_weights=use_focus_weights)
    # for site_name in sites:
    #     analyze_site(site_name,num_reals=num_reals,noptmax=noptmax,num_workers=num_workers,run=False,
    #         prep=True,plot=False,use_delay=True,use_focus_weights=use_focus_weights,include_ghb_pars=True,
    #         local=local)
    #     break
    # use_focus_weights = True
    # #try_analysis_for_sites(site_names,num_workers,num_reals,noptmax,use_focus_weights=use_focus_weights)
    # for site_name in sites:
    #     analyze_site(site_name,num_reals=num_reals,noptmax=noptmax,num_workers=num_workers,run=True,
    #         prep=True,plot=True,use_delay=True,use_focus_weights=use_focus_weights)
    # gather_pdfs()
    # gather_csvs(sites=['H201', 'GWM_14'])
    # gather_pdfs(sites=['H201', 'GWM_14'])
    prep_CH_scen_Input('EARLIMART')
        
        
    